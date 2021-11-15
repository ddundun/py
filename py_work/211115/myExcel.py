from openpyxl import *

class MyExcel():
    def savefn(self,kor,eng,math,tot,avg):
        wb= load_workbook("a.xlsx")
        ws = wb.active
        ws.append([kor,eng,math,tot,avg])
        wb.save('a.xlsx')
        wb.close()

    def loadfn(self):
        wb = load_workbook('a.xlsx')
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end = ' ')
            print()

    def createfn(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['국어','영어','수학','총점','평균'])
        wb.save('a.xlsx')
        wb.close()