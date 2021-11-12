from openpyxl import *
# workbook(생성), load_workbook(불러오기)
# wb.save('ddd.xlsx')/ wb.close()

class qExcel:
    def __init__(self):
        pass

    def save(self,a,b,c,d,e):
        wb= load_workbook('a.xlsx')
        ws = wb.active
        ws.append(["국어", "영어", "수학","평균", "총점"])
        ws.append([a,b,c,d,e])
        # ws.append([a,b,c,d,e])
        wb.save('a.xlsx')
        wb.close()

    def load(self):
        wb = load_workbook('a.xlsx')
        ws = wb.active

        # iter_rows : 행을 기준으로 반복하는 함수
        # iter_cols : 열을 기준으로 반복하는 함수

        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end = ' ')
            print()

        wb.save('a.xlsx')
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save('a.xlsx')


        wb.close()