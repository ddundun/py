from openpyxl import load_workbook, Workbook

class ScoreExcel:

    # 생성자
    def __init__(self):
        print("생성자")

    # 삭제..
    def deleterow(self,rowIndex):
        wb = load_workbook('score.xlsx')
        ws = wb.active

        # 엑셀에서는 1, 배열에선 0
        ws.delete_rows(rowIndex - 1)
        wb.save('score.xlsx')
        wb.close()

    # 불러오기
    def loadrow(self):
        rows = []
        try:
            wb = load_workbook('score.xlsx')

            ws = wb.active

            for row in ws.iter_rows(min_row=2): #2번째 줄 부터
                rows.append([row[0].value,row[1].value,row[2].value,row[3].value,row[4].value]) #1차원배열추가
            print(rows)

            wb.close()
            return rows
        except Exception as e :
            print(e)


    # 추가..
    def appendrow(self,kor,eng,math,tot,avg):
        wb = load_workbook("score.xlsx")
        ws = wb.active
        ws.append([kor, eng, math, tot, avg])
        wb.save('score.xlsx')
        wb.close()

    def createfile(self):
        wb =Workbook()
        ws = wb.active

        ws.append(['국어', '영어', '수학', '총점', '평균']) #첫번째 시트에 행넣기
        
        wb.save("score.xlsx")
        wb.close()

