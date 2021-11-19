from openpyxl import *
# workbook(생성), load_workbook(불러오기)
# wb.save('ddd.xlsx')/ wb.close()

class MExcel:
    def __init__(self):
        pass

    def save(self,a,b,c):
        wb= load_workbook('mexcel.xlsx')
        ws = wb.active
        ws.append([a,b,c])
        wb.save('mexcel.xlsx')
        wb.close()

    def load(self):
        wb = load_workbook('mexcel.xlsx')
        ws = wb.active

        # iter_rows : 행을 기준으로 반복하는 함수
        # iter_cols : 열을 기준으로 반복하는 함수
        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end = ' ')
            print()

        wb.save('mexcel.xlsx')
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save('mexcel.xlsx')
        wb.close()