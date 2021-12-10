from openpyxl import load_workbook, Workbook


class ProjectExcel:
    def loadrow(self):
        rows = []
        wb = load_workbook("member.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            rows.append([row[0].value,
                         row[1].value,
                         row[2].value])
        wb.close()
        return rows

    def appendrow(self, checkbox, name, number):
        wb = load_workbook("member.xlsx")
        ws = wb.active

        ws.append([checkbox, name, number])
        wb.save("member.xlsx")
        wb.close()

    def createfile(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['선택', '이름', '전화번호'])
        wb.save("member.xlsx")
        wb.close()

    def oneside(self):
        rows2 = []
        wb = load_workbook("member.xlsx")
        ws = wb.active
        for row2 in ws.iter_rows(min_row=2):
            rows2.append(row2[2].value)
        wb.close()
        return rows2
