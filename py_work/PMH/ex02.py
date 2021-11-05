from openpyxl import Workbook

wb= Workbook()

ws1= wb.create_sheet("newSheet")
ws2 = wb.create_sheet("NewSheet", 2)

ws1.sheet_properties.tabColor="9532a8"
ws2.sheet_properties.tabColor="457a87"

# a= {"aa":10,"bb":20}
#
# print(a["aa"])
# print(a["bb"])
# a["cc"]=40
# print(a)
for i in range(1,10):
    ws1['A'+str(i)] = 'test'+str(i)

ws2['B1'] ='ws2222'

target = wb.copy_worksheet(ws1)
target.title='copied sheet'

wb.save('b.xlsx')
wb.close()